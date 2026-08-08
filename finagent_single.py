"""finagent_single.py — the whole Financial PDF Extraction Agent in one file.

This is a portable, single-file build of the `finagent/` package. Same code,
flattened into one module so you can drop it anywhere, paste it in a chat, or
run it without installing a package.

    python finagent_single.py test_pdfs/TCS_2024-2025.pdf

The multi-file package in finagent/ is the source of truth.
AUTO-GENERATED FILE. Do not edit directly; modify finagent/ and run:
    python -m finagent.bundler
"""
import re
import sys
import time
from collections import defaultdict
from dataclasses import dataclass, field
from enum import Enum
from pathlib import Path
from statistics import median
from typing import Optional, List, Dict, Tuple

from pypdf import PdfReader
import pdfplumber
from rapidfuzz import fuzz
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

# Alias modular namespaces to the current module to support the bundled single-file build
profiler = sys.modules[__name__]
locator = sys.modules[__name__]
normalizer = sys.modules[__name__]
unit_detector = sys.modules[__name__]
geometric = sys.modules[__name__]
geometry = sys.modules[__name__]
tracing = sys.modules[__name__]


# =============================================================================

# SCHEMA (from finagent/schema.py)

# =============================================================================

# statement codes: PL = profit & loss, BS = balance sheet, CF = cash flow
METRICS = {
    # ---------------- Profit & Loss ----------------
    "revenue": {
        "statement": "PL",
        # "interest earned" is the bank revenue line (Schedule III P&L)
        "synonyms": [
            "revenue from operations",
            "net sales",
            "income from operations",
            "total revenues",
            "revenue",
            "sales",
            "turnover",
            "interest earned",
        ],
    },
    "other_income": {
        "statement": "PL",
        "synonyms": ["other income", "other operating income"],
    },
    "total_income": {
        "statement": "PL",
        "synonyms": ["total income", "total revenue"],
    },
    "total_expenses": {
        "statement": "PL",
        "synonyms": ["total expenses", "total expenditure", "total costs and expenses"],
    },
    "employee_costs": {
        "statement": "PL",
        "synonyms": [
            "employee benefit expenses",
            "employee benefits expense",
            "staff costs",
            "personnel expenses",
        ],
    },
    "depreciation": {
        "statement": "PL",
        # also matchable on the Cash Flow statement: function-of-expense P&Ls
        # (Newgen, HDFC) never print depreciation on the P&L — it only surfaces
        # as the first non-cash add-back in the CF reconciliation. The CF
        # wording ("depreciation of/on ...") is added so it is found there too.
        "also_on": ["CF"],
        "synonyms": [
            "depreciation and amortisation expense",
            "depreciation and amortization",
            "depreciation, amortisation and impairment",
            "depreciation amortisation and depletion expense",
            "depreciation of plant and equipment",
            "depreciation on fixed assets",
            "depreciation and amortisation",
        ],
    },
    "finance_costs": {
        "statement": "PL",
        # also on CF: same rationale as depreciation — appears as an add-back in
        # the operating-activities reconciliation when the P&L groups it away
        "also_on": ["CF"],
        "synonyms": ["finance costs", "finance cost", "interest expense", "borrowing costs"],
    },
    "profit_before_tax": {
        "statement": "PL",
        "synonyms": [
            "profit before tax",
            "profit before income tax",
            "income before income taxes",
            "profit/(loss) before tax",
            "earnings before tax",
        ],
    },
    "tax_expense": {
        "statement": "PL",
        "synonyms": [
            "total tax expense",
            "tax expense",
            "income tax expense",
            "provision for taxation",
            "income taxes",
        ],
    },
    "net_profit": {
        "statement": "PL",
        # bank P&L wording: the before-minority-interest line is the total
        # profit (incl. NCI), matching the "profit for the year" convention
        "synonyms": [
            "profit for the year",
            "net profit",
            "profit after tax",
            "net income",
            "profit/(loss) for the year",
            "profit for the period",
            "net profit/loss",
            "net profit for the year before minority interest",
            "profit after tax for the year",
        ],
    },
    "total_comprehensive_income": {
        "statement": "PL",
        "synonyms": ["total comprehensive income for the year", "total comprehensive income"],
    },
    "eps_basic": {
        "statement": "PL",
        "synonyms": [
            "basic earnings per share",
            "basic eps",
            "earnings per share basic",
            "basic (in rupees)",
            "basic (rs.)",
            "earnings per equity share basic",
            # combined line when basic == diluted (TCS); EU wording (BMW)
            "earnings per equity share basic and diluted",
            "basic earnings per ordinary share in eur",
            # power-sector wording (rate-regulated entities), as
            # truncated by the line wrap; the "before net movement"
            # twin is vetoed by the before/after directional guard
            "basic / diluted earnings per equity share after net movement",
        ],
    },
    # ---------------- Balance Sheet ----------------
    "total_assets": {
        "statement": "BS",
        "synonyms": ["total assets"],
    },
    "non_current_assets": {
        "statement": "BS",
        "synonyms": ["total non-current assets", "total non current assets", "non-current assets"],
    },
    "current_assets": {
        "statement": "BS",
        "synonyms": ["total current assets", "current assets"],
    },
    # side: these labels recur verbatim under BOTH BS sections; the metric
    # means the CURRENT-side row (extractor stamps rows with their section)
    "inventories": {
        "statement": "BS",
        "side": "current",
        "synonyms": ["inventories", "inventory"],
    },
    "trade_receivables": {
        "statement": "BS",
        "side": "current",
        "synonyms": ["trade receivables", "accounts receivable", "sundry debtors"],
    },
    "cash_and_equivalents": {
        "statement": "BS",
        "side": "current",
        "synonyms": ["cash and cash equivalents", "cash and bank balances"],
    },
    "total_equity": {
        "statement": "BS",
        "synonyms": [
            "total equity",
            "total shareholders' equity",
            "shareholders' funds",
            "equity attributable to owners",
            "equity",
        ],
    },
    "non_current_liabilities": {
        "statement": "BS",
        "synonyms": [
            "total non-current liabilities",
            "total non current liabilities",
            "non-current provisions and liabilities",
        ],
    },
    "current_liabilities": {
        "statement": "BS",
        "synonyms": ["total current liabilities", "current provisions and liabilities"],
    },
    "total_liabilities": {
        "statement": "BS",
        "synonyms": ["total liabilities"],
    },
    "total_equity_and_liabilities": {
        "statement": "BS",
        # "capital and liabilities" is the bank BS section (Schedule III)
        "synonyms": [
            "total equity and liabilities",
            "total liabilities and equity",
            "total liabilities and shareholders' equity",
            "total capital and liabilities",
        ],
    },
    # ---------------- Cash Flow ----------------
    "cash_from_operations": {
        "statement": "CF",
        # NOTE: "cash generated from operations" is deliberately absent —
        # it's the PRE-TAX subtotal (golden check caught it on Reliance,
        # Adani and Newgen); only "net …" totals belong here
        "synonyms": [
            "net cash generated from operating activities",
            "net cash from operating activities",
            "net cash flow from operating activities",
            "cash inflow/outflow from operating activities",
        ],
    },
    "cash_from_investing": {
        "statement": "CF",
        "synonyms": [
            "net cash used in investing activities",
            "net cash from investing activities",
            "net cash flow from investing activities",
            "cash inflow/outflow from investing activities",
            "net cash flows generated from investing activities",
            "net cash generated from investing activities",
        ],
    },
    "cash_from_financing": {
        "statement": "CF",
        "synonyms": [
            "net cash used in financing activities",
            "net cash from financing activities",
            "net cash flow from financing activities",
            "cash inflow/outflow from financing activities",
            "net cash flows generated from financing activities",
            "net cash generated from financing activities",
        ],
    },
    "net_change_in_cash": {
        "statement": "CF",
        "synonyms": [
            "net increase/(decrease) in cash and cash equivalents",
            "net increase in cash and cash equivalents",
            "net change in cash and cash equivalents",
            "net decrease in cash and cash equivalents",
            "net increase in cash and cash equivalents during the year",
        ],
    },
    "fx_effect_on_cash": {
        "statement": "CF",
        "optional": True,  # only multinationals report it; feeds the
        # cash-flow identity, never counted as MISSING
        "synonyms": [
            "effect of exchange rate on cash and cash equivalents",
            "effect of exchange rate changes on cash and cash equivalents",
            "effects of exchange rate changes on cash and cash equivalents",
            "exchange differences on cash and cash equivalents",
            "effect of exchange differences on cash and cash equivalents",
            "effect of fluctuation in foreign currency translation reserve",
            "exchange difference on translation of foreign currency cash and cash equivalents",
        ],
    },
    "closing_cash": {
        "statement": "CF",
        "synonyms": [
            "cash and cash equivalents at the end of the year",
            "cash and cash equivalents at end of the period",
            "closing cash and cash equivalents",
            "cash and cash equivalents as at the end of the year",
            "cash and cash equivalents as at december",
            "closing balance of cash and cash equivalents",
        ],
    },
}

ALL_METRICS = list(METRICS)
EXPECTED_METRICS = [m for m, d in METRICS.items() if not d.get("optional")]


def metrics_for_statement(code):
    """Metrics that may be matched on a statement: those that primarily live
    there, plus any whose `also_on` lists it (a metric that legitimately
    appears on more than one statement, e.g. depreciation on both P&L and CF)."""
    return [m for m, d in METRICS.items() if d["statement"] == code or code in d.get("also_on", [])]


# =============================================================================

# UNIT DETECTOR (from finagent/unit_detector.py)

# =============================================================================

@dataclass
class UnitInfo:
    unit_name: str  # e.g., "Crores", "Lakhs", "Millions", "Thousands", "Units"
    currency: str  # e.g., "INR", "USD", "EUR", "UNKNOWN"
    multiplier: float  # e.g., Crores -> 10_000_000, Lakhs -> 100_000, Millions -> 1_000_000
    raw_text: str


@dataclass
class PeriodHeader:
    current_period: str  # e.g. "31-Mar-2025" or "FY25"
    prior_period: str | None = None  # e.g. "31-Mar-2024" or "FY24"
    columns_detected: list[str] = None


UNIT_PATTERNS: list[tuple[str, str, str, float]] = [
    # (regex_pattern, unit_name, currency, multiplier)
    (r"(?:₹|rs\.?|rupees?)\s+(?:in\s+)?crores?", "Crores", "INR", 10_000_000.0),
    (r"(?:₹|rs\.?|rupees?)\s+(?:in\s+)?lakhs?", "Lakhs", "INR", 100_000.0),
    (r"(?:₹|rs\.?|rupees?)\s+(?:in\s+)?millions?", "Millions", "INR", 1_000_000.0),
    (r"(?:₹|rs\.?|rupees?)\s+(?:in\s+)?thousands?", "Thousands", "INR", 1_000.0),
    (r"usd\s+(?:in\s+)?millions?|\$\s+(?:in\s+)?millions?", "Millions", "USD", 1_000_000.0),
    (r"usd\s+(?:in\s+)?thousands?|\$\s+(?:in\s+)?thousands?", "Thousands", "USD", 1_000.0),
    (r"eur|€\s+(?:in\s+)?millions?", "Millions", "EUR", 1_000_000.0),
    (r"in\s+crores?", "Crores", "INR", 10_000_000.0),
    (r"in\s+lakhs?", "Lakhs", "INR", 100_000.0),
    (r"in\s+millions?", "Millions", "UNKNOWN", 1_000_000.0),
    (r"in\s+thousands?", "Thousands", "UNKNOWN", 1_000.0),
]

PERIOD_DATE_PATTERN = re.compile(
    r"(?:as\s+at|as\s+of|for\s+the\s+year\s+ended)?\s*"
    r"(\d{1,2}(?:st|nd|rd|th)?\s+(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+\d{4}"
    r"|\d{4}-\d{2}-\d{2}|fy\s*\d{2,4})",
    re.IGNORECASE,
)


def detect_unit(text: str) -> UnitInfo:
    """Scan page or table header text for unit scale declarations."""
    t = text.lower()
    for pattern, name, curr, mult in UNIT_PATTERNS:
        match = re.search(pattern, t)
        if match:
            return UnitInfo(unit_name=name, currency=curr, multiplier=mult, raw_text=match.group(0))

    return UnitInfo(unit_name="Units", currency="UNKNOWN", multiplier=1.0, raw_text="default")


def detect_periods(text: str) -> PeriodHeader:
    """Extract comparative period headers (e.g. 31 March 2025 vs 31 March 2024)."""
    matches = PERIOD_DATE_PATTERN.findall(text)
    if not matches:
        return PeriodHeader(current_period="Current", prior_period="Prior", columns_detected=[])

    unique_dates = []
    for m in matches:
        m_str = m.strip()
        if m_str not in unique_dates:
            unique_dates.append(m_str)

    current = unique_dates[0] if len(unique_dates) > 0 else "Current"
    prior = unique_dates[1] if len(unique_dates) > 1 else None
    return PeriodHeader(current_period=current, prior_period=prior, columns_detected=unique_dates)


# =============================================================================

# PROFILER (from finagent/profiler.py)

# =============================================================================

@dataclass
class PageProfile:
    index: int  # 0-based
    width: float
    height: float
    landscape: bool
    text: str  # raw extracted text (kept for the locator)
    text_quality: str  # OK / SUSPECT / EMPTY


@dataclass
class DocProfile:
    path: str
    n_pages: int
    pages: list = field(default_factory=list)
    logical_pages: list = field(default_factory=list)  # NEW: logical pages after geometry split

    @property
    def landscape_ratio(self):
        return sum(p.landscape for p in self.pages) / max(self.n_pages, 1)

    def summary(self):
        q = {}
        for p in self.pages:
            q[p.text_quality] = q.get(p.text_quality, 0) + 1
        return {
            "pages": self.n_pages,
            "landscape_ratio": round(self.landscape_ratio, 2),
            "text_quality": q,
        }


def _quality(text):
    if not text or len(text.strip()) < 50:
        return "EMPTY"
    alpha = sum(len(w) for w in re.findall(r"[A-Za-z]{3,}", text))
    return "OK" if alpha / max(len(text), 1) > 0.3 else "SUSPECT"


def profile(pdf_path):
    reader = PdfReader(pdf_path)
    doc = DocProfile(path=str(pdf_path), n_pages=len(reader.pages))
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(reader.pages):
            box = page.mediabox
            w, h = float(box.width), float(box.height)
            rotation = page.get("/Rotate") or 0
            if rotation in (90, 270):
                w, h = h, w
            try:
                text = page.extract_text() or ""
            except Exception:  # noqa: BLE001
                text = ""
            doc.pages.append(
                PageProfile(
                    index=i,
                    width=w,
                    height=h,
                    landscape=w > h,
                    text=text,
                    text_quality=_quality(text),
                )
            )

            # --- GEOMETRY STAGE: Split physical page into logical pages ---
            pdfplumber_page = pdf.pages[i]
            words = pdfplumber_page.extract_words(use_text_flow=False, keep_blank_chars=False)
            upright = [w for w in words if w.get("upright", True)]
            if len(upright) >= len(words) / 2:
                words = upright
            logical = geometry.logical_pages(pdfplumber_page, words)
            for group in logical:
                # Extract text from the logical page (for scoring)
                logical_text = " ".join(w["text"] for w in group)
                doc.logical_pages.append(
                    {
                        "physical_page": i,
                        "text": logical_text,
                        "text_quality": _quality(logical_text),
                    }
                )
    return doc


# =============================================================================

# GEOMETRY (from finagent/geometry.py)

# =============================================================================

GUTTER_BINS = 200  # x-axis resolution for the coverage histogram
SEARCH_LO, SEARCH_HI = 0.35, 0.65  # look for the gutter in the middle third
MAX_GUTTER_COVERAGE = 0.01  # words allowed to touch the gutter band
MIN_SIDE_RATIO = 0.2  # both halves must carry real text


def _gutter_x(page_width, words):
    """X position of the spine: the near-empty vertical band CLOSEST TO THE
    CENTRE (not the emptiest one — a table's gap between its label column
    and its number columns is often emptier than the true spine)."""
    cover = [0] * GUTTER_BINS
    for w in words:
        a = max(int(w["x0"] / page_width * GUTTER_BINS), 0)
        b = min(int(w["x1"] / page_width * GUTTER_BINS), GUTTER_BINS - 1)
        for i in range(a, b + 1):
            cover[i] += 1
    lo, hi = int(GUTTER_BINS * SEARCH_LO), int(GUTTER_BINS * SEARCH_HI)
    threshold = max(1, MAX_GUTTER_COVERAGE * len(words))
    candidates = [i for i in range(lo, hi) if cover[i] <= threshold]
    if not candidates:
        return None
    centre = GUTTER_BINS / 2
    best = min(candidates, key=lambda i: abs(i + 0.5 - centre))
    return (best + 0.5) / GUTTER_BINS * page_width


def split_two_up(page, words):
    """Return (left_words, right_words) if the page is two-up, else None."""
    if page.width <= page.height or len(words) < 20:
        return None
    gx = _gutter_x(page.width, words)
    if gx is None:
        return None
    left = [w for w in words if (w["x0"] + w["x1"]) / 2 <= gx]
    right = [w for w in words if (w["x0"] + w["x1"]) / 2 > gx]
    if min(len(left), len(right)) < MIN_SIDE_RATIO * max(len(left), len(right), 1):
        return None  # one side nearly empty: a wide table, not two pages
    for side in (left, right):
        alpha = sum(1 for w in side if any(c.isalpha() for c in w["text"]))
        if alpha < 0.3 * len(side):
            return None  # a side that is mostly numbers is a value-column
            # block of one wide table, not a logical page
    return left, right


def logical_pages(page, words):
    """One physical page -> list of logical word groups (1 or 2)."""
    halves = split_two_up(page, words)
    return list(halves) if halves else [words]


# =============================================================================

# LOCATOR (from finagent/locator.py)

# =============================================================================

# What counts as a "money figure" when deciding a page carries a real
# statement (not a ToC mention). Two shapes: a long comma-grouped integer
# ("1,58,788") OR a decimal figure with two places ("8207.65"). The decimal
# shape is essential for reports printed in crore, where consolidated line
# items are small 4-digit decimals that the comma-grouped pattern misses
# entirely — BHEL's whole P&L scored 0 on the integer-only count.
NUM_RE = re.compile(r"\d[\d,]{4,}|\d[\d,]*\.\d{2}\b")

# (title patterns, content cue patterns) — title hits score heavily,
# cues confirm we're on the statement itself rather than a ToC mention.
STATEMENT_SIGNATURES = {
    "BS": (
        [r"balance sheet", r"statement of financial position"],
        # second row: bank wording (Banking Regulation Act Schedule III) —
        # banks print Capital and Liabilities / Deposits / Advances with no
        # current/non-current split, so corporate cues never fire
        [
            r"total assets",
            r"equity and liabilities",
            r"total equity",
            r"non[- ]current assets",
            r"current liabilities",
            r"capital and liabilities",
            r"reserves and surplus",
            r"\bdeposits\b",
            r"\badvances\b",
            r"\bborrowings\b",
        ],
    ),
    "PL": (
        # "profit AND loss" is Indian GAAP wording; IFRS reports (Singapore,
        # EU) title the same statement "profit OR loss" / "comprehensive income"
        [
            r"statement of profit (?:and|or) loss",
            r"income statement",
            r"statement of (?:operations|income)",
            r"profit and loss account",
            r"statement of comprehensive income",
        ],
        [
            r"revenue from operations",
            r"total income",
            r"profit before tax",
            r"earnings per (?:equity )?share",
            r"total expenses",
            r"gross profit",
            r"cost of sales",
            r"profit for the year",
            r"income tax expense",
        ],
    ),
    "CF": (
        [r"(?:statement of )?cash flows?", r"cash flow statement"],
        [r"operating activities", r"investing activities", r"financing activities"],
    ),
}


@dataclass
class Location:
    statement: str  # BS / PL / CF
    basis: str  # consolidated / standalone / unknown
    page_indices: list  # 0-based, best first
    score: float


def _search(pat, text):
    """re.search with a kerning-tolerant fallback: PDF text layers sometimes
    split a word internally ("BAL ANCE SHEET"), so retry with the pattern's
    literal spaces removed against space-collapsed text."""
    return re.search(pat, text) or re.search(pat.replace(" ", ""), text.replace(" ", ""))


def _is_heading(line, title_pats):
    """True if `line` is a statement HEADING (the title phrase starts at its
    head, after at most a 2-word basis/company prefix) rather than a prose
    mention that merely contains the phrase. Kerning-split titles ("BAL ANCE
    SHEET") count as headings — they only arise on real heading lines."""
    for p in title_pats:
        m = re.search(p, line)
        if m:
            if len(line[: m.start()].split()) <= 2:
                return True
        elif re.search(p.replace(" ", ""), line.replace(" ", "")):
            return True
    return False


def verify_page(pdf_path: str, page_num: int, cue_pats: list) -> bool:
    """Verify a candidate page has:
    - >= 8 numeric tokens (already checked by locator)
    - >= 2 distinct x-aligned numeric columns (catches prose/notes pages)
    - >= 5 raw line items from the geometric extractor (fast dry-run)
    """
    with pdfplumber.open(pdf_path) as pdf:
        page = pdf.pages[page_num - 1]  # 1-indexed
        text = page.extract_text()

        # (A) Token check (already in locator, but double-check)
        numeric_tokens = len(NUM_RE.findall(text))
        if numeric_tokens < 8:
            return False

        # (B) Column check: >= 2 distinct x-aligned numeric columns
        words = page.extract_words()
        x_coords = [w["x0"] for w in words if w["text"].replace(".", "").isdigit()]
        distinct_columns = len({round(x / 50) * 50 for x in x_coords})  # 50px tolerance
        if distinct_columns < 2:
            return False

        # (C) Dry-run extractor: >= 5 raw line items
        try:
            items = extract(pdf_path, [page_num - 1], cue_pats)  # 0-based
            if len(items) < 5:
                return False
        except Exception:  # noqa: BLE001
            return False

    return True


def _score_page(text, title_pats, cue_pats):
    t = text.lower()
    title = sum(3 for p in title_pats if _search(p, t))
    cues = sum(1 for p in cue_pats if re.search(p, t))
    if title == 0 or cues < 2:
        return 0, "unknown", False
    # ToC pages mention many statements but have few numbers
    numbers = len(NUM_RE.findall(t))
    if numbers < 8:
        return 0, "unknown", False
    # a real statement carries its title as a top-of-page heading; commentary
    # pages (management report) merely mention the statement in prose
    head_lines = [ln.strip() for ln in t.strip().splitlines()[:6] if ln.strip()]
    # a TITLE is a heading line whose statement phrase sits at the HEAD of the
    # line, optionally after a short basis/company prefix ("Standalone Balance
    # Sheet", "BMW AG Balance Sheet", "Balance Sheet for Group and Segments at
    # 31 December 2025"). A prose sentence buries the phrase mid-line
    # ("Provisions are reviewed at each balance sheet date ..." on a notes
    # page) — that is NOT a title. Without this guard a notes sentence becomes
    # a fake heading, takes the +5 boost and a basis stamp, and outranks the
    # real statement. Length alone fails (dated titles run to ~10 words);
    # position is the discriminator: phrase must start within the first 2 words.
    title_line = next((ln for ln in head_lines if _is_heading(ln, title_pats)), None)
    # "Condensed/Summarised X" headings are management-report summaries,
    # not the statement itself — no heading boost, no basis authority
    if title_line and re.search(r"condensed|summaris|summariz|abridged", title_line):
        title_line = None
    # real statements show two comparative periods (IAS 1); a page parading
    # many distinct years is a ten-year/multi-year overview. Heading-titled
    # pages get slack — footnotes legitimately cite years (Airtel's spectrum
    # auction list) — pages without a true heading do not
    years = {m for m in re.findall(r"\b(?:19|20)\d{2}\b", t)}
    if len(years) >= (10 if title_line else 6):
        return 0, "unknown", False
    if title_line:
        title += 5
    # the title line declares the basis ("Consolidated Balance Sheet",
    # German "for Group and Segments"); whole-page text is only a fallback —
    # running headers/prose mentioning "group"/"consolidated" must not count
    basis = "unknown"
    if title_line and re.search(r"consolidated|\bgroup\b", title_line):
        basis = "consolidated"
    elif title_line and re.search(r"standalone|\bseparate\b", title_line):
        basis = "standalone"
    elif "consolidated" in t:
        basis = "consolidated"
    elif "standalone" in t or "separate financial" in t:
        basis = "standalone"
    return title + cues + min(numbers, 30) * 0.1, basis, bool(title_line)


def _scored_pages(doc_profile, title_pats, cue_pats):
    """All logical pages that score as this statement: (score, basis, heading, logical_index)."""
    scored = []
    for logical_idx, logical_page in enumerate(doc_profile.logical_pages):
        if logical_page["text_quality"] == "EMPTY":
            continue
        s, basis, heading = _score_page(logical_page["text"], title_pats, cue_pats)
        if s > 0:
            scored.append((s, basis, heading, logical_idx))
    return scored


def _pick(
    scored, doc_profile, cue_pats, code, want_basis=None, prefer_consolidated=False, exclude=()
):
    """Choose the best statement page (+ its continuations) from `scored`.

    want_basis: if given, restrict to pages stamped that basis (used to find
    the standalone counterpart). prefer_consolidated: soft preference within
    the tier (the primary selection). exclude: logical page indices already taken by
    the primary, so the alternate basis can't reuse them.
    """
    if not scored:
        return None
    # heading-titled pages are real statements; pages that merely discuss the
    # statement (notes, management report) rank below them — a notes page
    # saying "consolidated" must not hijack the basis pool
    headed = [x for x in scored if x[2]]
    pool = headed or scored
    pool = [x for x in pool if x[3] not in exclude]
    if want_basis is not None:
        pool = [x for x in pool if x[1] == want_basis]
    elif prefer_consolidated:
        consolidated = [x for x in pool if x[1] == "consolidated"]
        pool = consolidated or pool
    if not pool:
        return None
    # ties go to the EARLIER page: the statement itself precedes the
    # notes/SOCIE pages that echo its keywords
    pool.sort(key=lambda x: (-x[0], x[3]))

    # --- VERIFIED PAGE SELECTION ---
    # Try each candidate in score order until one passes verification
    for score, basis, _, best in pool:
        # Get the physical page and logical text for verification
        logical_page = doc_profile.logical_pages[best]
        physical_page = logical_page["physical_page"] + 1  # 1-indexed
        if verify_page(doc_profile.path, physical_page, cue_pats):
            break
    else:
        # No candidate passed verification; fall back to the top-scoring one
        score, basis, _, best = pool[0]
        logical_page = doc_profile.logical_pages[best]
        physical_page = logical_page["physical_page"] + 1  # 1-indexed

    # Statements may continue on a neighbouring logical page (which lacks the title
    # there). Only adjacent logical pages on the SAME physical page qualify.
    logical_page = doc_profile.logical_pages[best]
    physical_page = logical_page["physical_page"]
    pages = [physical_page]  # Return 0-indexed physical page
    for nb in (best - 1, best + 1):
        if 0 <= nb < len(doc_profile.logical_pages):
            neighbour = doc_profile.logical_pages[nb]
            if neighbour["physical_page"] == physical_page and _is_continuation(
                neighbour["text"], cue_pats
            ):
                pages.append(physical_page)  # Continuation on same physical page
    return Location(code, basis, sorted(set(pages)), score)


def locate(doc_profile):
    """Return {statement_code: Location} for BS, PL, CF — the PRIMARY selection
    (prefers consolidated, falls back per-statement). Unchanged behaviour."""
    results = {}
    for code, (title_pats, cue_pats) in STATEMENT_SIGNATURES.items():
        scored = _scored_pages(doc_profile, title_pats, cue_pats)
        loc = _pick(scored, doc_profile, cue_pats, code, prefer_consolidated=True)
        results[code] = loc or Location(code, "unknown", [], 0)
    return results


def locate_alternate(doc_profile, primary):
    """Find the OTHER basis's statement pages, distinct from `primary`.

    If the primary selection landed on consolidated pages, this returns the
    standalone counterpart (and vice-versa) so both can be extracted and shown
    side by side. Returns {code: Location}; a code with no counterpart gets an
    empty Location (basis "none"), which the pipeline reads as "this basis was
    not present in the PDF" rather than "extracted nothing".
    Supports fallback to the primary page if it contains combined multi-column tables.
    """
    results = {}
    for code, (title_pats, cue_pats) in STATEMENT_SIGNATURES.items():
        prim = primary.get(code)
        prim_basis = prim.basis if prim else "unknown"
        want = (
            "standalone"
            if prim_basis == "consolidated"
            else "consolidated"
            if prim_basis == "standalone"
            else None
        )  # primary basis unknown -> no distinct counterpart
        if want is None:
            results[code] = Location(code, "none", [], 0)
            continue
        scored = _scored_pages(doc_profile, title_pats, cue_pats)
        exclude = set(prim.page_indices) if prim else set()
        loc = _pick(scored, doc_profile, cue_pats, code, want_basis=want, exclude=exclude)
        # Fallback for combined tables (same page contains both bases in multi-column layout)
        if not (loc and loc.page_indices) and prim and prim.page_indices:
            for p_idx in prim.page_indices:
                if 0 <= p_idx < len(doc_profile.logical_pages):
                    txt = doc_profile.logical_pages[p_idx]["text"].lower()
                    if ("standalone" in txt or "separate" in txt) and (
                        "consolidated" in txt or "group" in txt
                    ):
                        loc = Location(code, want, prim.page_indices, prim.score)
                        break
        results[code] = loc or Location(code, want, [], 0)
    return results


def has_pages(locations):
    """True if any statement in this basis actually resolved to pages."""
    return any(loc.page_indices for loc in locations.values())


def _is_continuation(text, cue_pats):
    t = text.lower()
    cues = sum(1 for p in cue_pats if re.search(p, t))
    numbers = len(NUM_RE.findall(t))
    return cues >= 1 and numbers >= 8


# =============================================================================

# GEOMETRIC EXTRACTOR (from finagent/extractors/geometric.py)

# =============================================================================

# tokens that look like report numbers: 1,49,982.45  (1,234)  123.45  -
NUM_CHARS = set("0123456789,().%-−–—")

# European reports render negatives as a DETACHED minus: "− 112,858" is two
# words. Only unicode minuses get merged — a standalone ASCII "-" is a nil
# placeholder in Indian reports and must stay a separate token.
MINUS_TOKENS = {"−", "–", "—"}

# a real sign hugs its number; a nil placeholder ("−" meaning zero) sits in
# its own column, far from the next number
MAX_SIGN_GAP = 4.0

# standalone dashes are nil values: part of a row's numeric tail
PLACEHOLDER_TOKENS = {"-", "−", "–", "—"}

# A 4-digit period header ("2025", "(2024)") used to confirm we found the
# column-header row and to count the value columns.
YEAR_RE = re.compile(r"\(?(?:19|20)\d{2}\)?$")
# the reference-column header word ("Note", "Schedule", "Page" ...) — the
# value columns sit to its right, so its right edge is the column boundary.
REF_HEADER_RE = re.compile(r"^(?:notes?|schedule|sch|page|ref)\.?$", re.IGNORECASE)
# small gap to the right of the reference header before the value columns
REF_FLOOR_MARGIN = 12.0
# more year columns than this means a segmented sheet (BMW: a 2025/2024 pair
# per business segment) or a multi-year overview. The value column is then no
# longer simply "the leftmost", so positional cutting is unsafe — bail out.
MAX_VALUE_COLS = 3


@dataclass
class RawItem:
    label: str
    values: list  # raw strings, left to right
    page: int  # 0-based physical page
    source: str = "geometric"
    side: str = None  # current / non-current, from the enclosing BS section


def _is_numeric_token(tok):
    t = tok.strip()
    if not t or not any(c.isdigit() for c in t):
        return False
    return all(c in NUM_CHARS for c in t)


def _merge_detached_minus(line):
    """x0-sorted words -> [(token, x0)], gluing a detached minus onto the
    number it signs. Distance decides sign vs nil placeholder. The x0 (left
    edge) rides along so a token can later be assigned to its column."""
    out, i = [], 0
    while i < len(line):
        w = line[i]
        if (
            w["text"] in MINUS_TOKENS
            and i + 1 < len(line)
            and _is_numeric_token(line[i + 1]["text"])
            and line[i + 1]["x0"] - w["x1"] <= MAX_SIGN_GAP
        ):
            out.append(("-" + line[i + 1]["text"], w["x0"]))
            i += 2
        else:
            out.append((w["text"], w["x0"]))
            i += 1
    return out


def _detect_value_floor(lines):
    """Locate the left boundary of the value columns.

    Returns an x-coordinate: any number whose left edge sits below it is a
    reference-column entry (note / schedule / page number), not a value.
    Returns None when no confident header is found — callers then fall back to
    the magnitude heuristics.

    Two anchors. First the period-header row (the line with the most 4-digit
    year tokens) tells us where the value columns START (leftmost year) and how
    MANY there are — too many means a segmented/multi-year sheet (BMW prints a
    2025/2024 pair per business segment) where the value column isn't simply
    "leftmost", so we bail. Then the boundary itself is taken from the
    reference-column header word ("Note"/"Schedule"/"Page") sitting left of the
    values: its right edge is exactly where the value columns begin. Anchoring
    on the ref header beats anchoring on the year token, which can drift far
    right inside a wide "As at March 31, 2025" header while the numbers below
    align further left (BHEL).
    """
    xs = [w["x0"] for ln in lines for w in ln]
    if not xs or max(xs) - min(xs) < 50:  # no real column spread
        return None
    best = None  # (year_count, leftmost_year_x0)
    for ln in lines:
        yrs = [w["x0"] for w in ln if YEAR_RE.fullmatch(w["text"])]
        if not yrs:
            continue
        cand = (len(yrs), min(yrs))  # most years wins; tie -> rightmost
        if best is None or cand > best:
            best = cand
    if best is None or best[0] > MAX_VALUE_COLS:
        return None
    anchor = best[1]  # leftmost value-year column
    # rightmost reference-column header that sits left of the value columns
    ref_x1 = None
    for ln in lines:
        for w in ln:
            if REF_HEADER_RE.match(w["text"]) and w["x1"] < anchor:
                ref_x1 = w["x1"] if ref_x1 is None else max(ref_x1, w["x1"])
    if ref_x1 is None:  # no reference column to strip
        return None
    return ref_x1 + REF_FLOOR_MARGIN


def _lines_from_words(words, y_tol=2.5):
    """Group words into visual lines by their vertical position."""
    lines = []
    for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
        if lines and abs(w["top"] - lines[-1][-1]["top"]) <= y_tol:
            lines[-1].append(w)
        else:
            lines.append([w])
    return lines


def _detect_column_bases(lines):
    """Scan header lines for column-spanning basis declarations ('Standalone' vs 'Consolidated').

    Returns a list of dicts: [{'basis': 'standalone', 'x0': float, 'x1': float}, ...] or empty list.
    """
    ranges = []
    for line in lines[:12]:
        text_line = " ".join(w["text"] for w in line).lower()
        if (
            "standalone" in text_line
            or "consolidated" in text_line
            or "separate" in text_line
            or "group" in text_line
        ):
            for w in line:
                t = w["text"].lower()
                if re.search(r"standalone|\bseparate\b", t):
                    ranges.append({"basis": "standalone", "x0": w["x0"], "x1": w["x1"] + 150})
                elif re.search(r"consolidated|\bgroup\b", t):
                    ranges.append({"basis": "consolidated", "x0": w["x0"], "x1": w["x1"] + 150})
    if not ranges:
        return []
    ranges.sort(key=lambda r: r["x0"])
    for i in range(len(ranges) - 1):
        mid = (ranges[i]["x1"] + ranges[i + 1]["x0"]) / 2
        ranges[i]["x1"] = mid
        ranges[i + 1]["x0"] = mid
    if ranges:
        ranges[0]["x0"] = 0.0
        ranges[-1]["x1"] = 9999.0
    return ranges


def _items_from_words(words, page_index, want_basis=None):
    items = []
    section = None
    side = None  # current / non-current: duplicate labels ("- Trade
    # receivables") appear under both BS sections
    side_heading = None  # text of the line that set the side; an
    # UNLABELED numeric row ENDING its block is the
    # section subtotal (Airtel prints "4,467,716
    # 3,862,549" with no label at all)
    pending = None  # candidate subtotal: only real if no labeled row
    # follows it before the block closes (a subtotal
    # is the LAST row of its section, so a mid-block
    # orphan from a wrapped label is discarded)
    prev_text = None  # immediately-preceding text-only line: a numeric
    # row whose label starts lowercase is its WRAPPED
    # continuation ("Profit for the year, representing
    # total" + "comprehensive income ... 419,056")

    def flush_pending():
        nonlocal pending
        if pending:
            items.append(pending)
            pending = None

    lines = _lines_from_words(words)
    # locate the value columns once for the whole block; None falls back to
    # the normalizer's magnitude heuristics (header not confidently found)
    value_floor = _detect_value_floor(lines)
    basis_ranges = _detect_column_bases(lines) if want_basis else []
    for line in lines:
        line.sort(key=lambda w: w["x0"])
        toks = _merge_detached_minus(line)  # [(text, x0), ...]
        texts = [t for t, _ in toks]
        # split into label prefix and trailing numeric tokens; nil
        # placeholders are tail members, not label text
        split = len(toks)
        for i in range(len(toks) - 1, -1, -1):
            if _is_numeric_token(texts[i]) or texts[i] in PLACEHOLDER_TOKENS:
                split = i
            else:
                break
        label = " ".join(texts[:split]).strip()
        num_pairs = toks[split:]
        # a parenthetical can split across the boundary: "... (refer note"
        # | "15)" — the closing fragment looks numeric and would become the
        # value (closing_cash = 15!). Re-join it into the label.
        while (
            label.count("(") > label.count(")")
            and num_pairs
            and re.fullmatch(r"[^()]*\)", num_pairs[0][0])
        ):
            label = f"{label} {num_pairs.pop(0)[0]}"
        # COLUMN GEOMETRY: a numeric token sitting LEFT of the leftmost value
        # column is a reference-column entry (note / schedule / page number),
        # not a value — drop it by POSITION, regardless of magnitude. This is
        # what lets a bare small integer ("5") under a year header be read as
        # a value. Guarded: if the filter would wipe out every number (header
        # mis-detected for this row), keep the row and let the magnitude
        # heuristics in the normalizer handle it.
        if value_floor is not None:
            kept = [p for p in num_pairs if p[1] >= value_floor]
            if kept:
                num_pairs = kept

        if want_basis and basis_ranges:
            matching = [r for r in basis_ranges if r["basis"] == want_basis]
            if matching:
                filtered = [
                    p for p in num_pairs if any(r["x0"] <= p[1] <= r["x1"] for r in matching)
                ]
                if filtered:
                    num_pairs = filtered

        nums = [t for t, _ in num_pairs]
        if label and nums:
            if prev_text and label[:1].islower():
                label = f"{prev_text} {label}"
            prev_text = None
            # bank/RBI-format statements label section totals just "Total"
            # (and EPS rows just "Basic"/"Diluted" under their heading);
            # qualify with the section heading ("ASSETS" -> "total assets")
            # so the normalizer can match it. A misattributed section simply
            # fails the fuzzy threshold and the row stays unmatched.
            if section and re.fullmatch(r"(total|basic|diluted):?", label, re.IGNORECASE):
                label = f"{label.rstrip(':')} {section}"
            # a total-ish row ("TOTAL LIABILITIES", "NET CURRENT ASSETS")
            # closes the block — totals FOLLOW subtotals, so a pending
            # unlabeled row right before it IS the subtotal (Wilmar).
            # Any other labeled row means the pending row wasn't the
            # block's last — a wrapped-label orphan, discarded.
            if re.match(r"(total|net)\b", label, re.IGNORECASE):
                flush_pending()
            else:
                pending = None
            if re.match(r"total\b", label, re.IGNORECASE):
                side_heading = None
            items.append(RawItem(label=label, values=nums, page=page_index, side=side))
        elif nums and side_heading:
            prev_text = None
            # unlabeled numeric row inside a current/non-current block:
            # candidate subtotal (last one wins). A wrong synthesis either
            # fails the fuzzy gate (absence), loses the first-wins tie to
            # the real row, or breaks a composition identity (FLAGGED).
            pending = RawItem(
                label=f"total {side_heading}", values=nums, page=page_index, side=side
            )
        elif label:
            flush_pending()  # a heading closes the block: the pending
            # unlabeled row WAS its last row = subtotal
            prev_text = label
            # a text-only line is a section heading candidate; strip leading
            # roman/decimal numbering ("I. INCOME" -> "income") and heading
            # parentheticals, which are unit/face-value noise ("EARNINGS PER
            # EQUITY SHARE (Face value 1/- per share)")
            section = re.sub(
                r"^[\divxlc]+[.)]\s*", "", re.sub(r"\([^)]*\)", " ", label.lower())
            ).strip()
            # "non-current" contains "current": test it first. Headings that
            # mention neither leave the side untouched ("Financial assets").
            if re.search(r"non[- ]current", section):
                side = "non-current"
                side_heading = section
            elif "current" in section:
                side = "current"
                side_heading = section
    flush_pending()
    return items


def _matches_statement(words, cue_pats):
    """Does this logical half belong to the statement we were sent to?

    Threshold is 1 cue, not 2: the locator already vouched for the page, so
    the half-picker only needs to drop the OTHER statement sharing an A3
    two-up sheet. A statement can SPAN both halves (Airtel CF: financing
    section alone on the right half, exactly 1 cue) — requiring 2 cues
    silently discarded it."""
    text = " ".join(w["text"] for w in words).lower()
    return sum(1 for p in cue_pats if re.search(p, text)) >= 1


def extract(pdf_path, page_indices, cue_pats=None, want_basis=None):
    """Extract raw line items from the given 0-based physical pages.

    cue_pats: content cues of the target statement. Used only to pick the
    right half of a split two-up page; if no half matches, keep all (the
    normalizer's allowed-metrics filter is the second line of defence).
    want_basis: optional 'standalone' or 'consolidated' filter for multi-column tables.
    """
    items = []
    with pdfplumber.open(pdf_path) as pdf:
        for idx in page_indices:
            if not (0 <= idx < len(pdf.pages)):
                continue
            page = pdf.pages[idx]
            words = page.extract_words(use_text_flow=False, keep_blank_chars=False)
            # rotated words are page furniture (vertical "Financial
            # Statements" tabs) that share a y-band with table rows and
            # break the numeric-tail scan. On a landscape page EVERYTHING
            # is rotated — there the filter must stand down.
            upright = [w for w in words if w.get("upright", True)]
            if len(upright) >= len(words) / 2:
                words = upright
            logical = geometry.logical_pages(page, words)
            if len(logical) > 1 and cue_pats:
                matching = [g for g in logical if _matches_statement(g, cue_pats)]
                logical = matching or logical
            for group in logical:
                items.extend(_items_from_words(group, idx, want_basis=want_basis))
    return items


# =============================================================================

# NORMALIZER (from finagent/normalizer.py)

# =============================================================================

MATCH_THRESHOLD = 88


@dataclass
class Extraction:
    metric: str
    value: float
    raw_label: str
    page: int
    source: str
    score: float
    extra_values: list  # remaining columns (usually prior year)


def parse_number(tok):
    """'(1,234.5)' -> -1234.5 ; '1,49,982' -> 149982.0 ; '-' -> None"""
    t = tok.strip().rstrip("*#†")
    t = t.replace("−", "-").replace("–", "-").replace("—", "-")
    if not t or not any(c.isdigit() for c in t):
        return None
    neg = t.startswith("(") and t.endswith(")")
    t = t.strip("()").replace(",", "").replace("%", "").strip()
    if t.startswith("-"):
        neg, t = True, t[1:]
    try:
        v = float(t)
    except ValueError:
        return None
    return -v if neg else v


def _looks_like_note_ref(tok, rest, label=""):
    """A bare small integer in the first numeric slot is a note-reference
    column, even when it is the ONLY number (a split table can leave the
    note ref behind while the value columns land elsewhere — taking it as
    the value would yield revenue=25). Real values carry commas, decimals,
    parens, or more digits."""
    # if the note ref was consumed into the label ("Inventories 10(e)"),
    # the first numeric token IS the value — dropping it took prior-year 28
    # over golden 21 (TCS)
    if re.search(r"\d[a-z()]*\s*$", label):
        return False
    t = tok.strip()
    if re.fullmatch(r"\d{1,2}", t):
        return True
    # a 3-digit bare int is usually a value, not a note number (Airtel fx
    # effect 718); treat as note ref only with the ~100x jump a real
    # label->note->value row shows
    if re.fullmatch(r"\d{3}", t) and rest:
        nxt = parse_number(rest[0])
        return nxt is not None and abs(nxt) >= 100 * float(t)
    # decimal note refs exist too (Adani numbers notes "5.5"); only a note
    # ref when the next value is orders of magnitude larger — an EPS of
    # 8.05 followed by prior-year 10.20 must survive
    if re.fullmatch(r"\d{1,2}\.\d{1,2}", t) and rest:
        nxt = parse_number(rest[0])
        return nxt is not None and abs(nxt) >= 100 * float(t)
    return False


def _looks_like_page_ref(tok, nxt):
    """A SECOND leading reference column. Some statements (BHEL) print both a
    Note column AND a Page cross-ref column before the values:
    "Inventories | 10 | 289 | 9,869.49". The note ref (10) is stripped first;
    289 is the page where the note lives, not a value. The 3-digit jump rule in
    _looks_like_note_ref can't catch it (289 -> 9,869 is only ~34x), so it would
    be taken as inventories=289. Safe signature: only fires once a note ref has
    already been stripped (single-reference tables never reach here), the token
    is a bare 1-3 digit integer, and a real money figure (comma or decimal)
    follows it — a page ref always precedes the value columns."""
    t = tok.strip()
    if not re.fullmatch(r"\d{1,3}", t):
        return False
    n = (nxt or "").strip()
    return bool(re.search(r"[.,]", n)) and parse_number(n) is not None


# sign/unit/reference qualifiers carry no identity — "/(loss)", "/loss",
# "(used in)", "(decrease)", "(in rupees)", "(note 24)", "(net of refunds)" —
# and are stripped from labels AND synonyms alike. Classification qualifiers
# ("(non-current)", "(current)") DO carry identity and must survive: stripping
# them once let the non-current line shadow the real value (TCS/Adani/Airtel
# golden WRONGs), because both then clean to the same string and the
# non-current section comes first.
# bare "(net)" is NOT a qualifier here: stripping it turned "Current tax
# liabilities (net)" into a 92-score match for total current liabilities
_QUALIFIER = (
    r"(?:loss(?:es)?|decrease|used in|in rupees|rs\.?|"
    r"net of [^)]*|refer[^)]*|notes?\s*[\d., ]*|continued|"
    r"face value[^)]*|"  # (Face Value of Rs 10 each)
    r"[a-z](?:\s*\+\s*[a-z])*)"
)  # cross-refs: (a), (a+b+c)


def clean_label(label):
    t = label.lower()
    t = re.sub(rf"/\s*\(?{_QUALIFIER}\)?(?=[\s)]|$)", " ", t)  # /(loss), /loss
    t = re.sub(rf"\(\s*{_QUALIFIER}\s*\)", " ", t)  # (used in)
    t = re.sub(r"^[\divxlc]+[.)]\s*", "", t)  # leading numbering: 1. / (iv)
    # fully-parenthesized enumerators: "(ii) Trade Receivables", "(a) …"
    t = re.sub(r"^\((?:[ivxlc]{1,4}|[a-z]|\d{1,2})\)\s*", "", t)
    # a LEADING basis word is page-level info the locator already resolved
    # ("Consolidated Net Profit for the year..."), not label identity
    t = re.sub(r"^(?:consolidated|standalone)\s+", "", t)
    t = re.sub(r"[^a-z()/'& -]", " ", t)  # drop stray digits/symbols
    return re.sub(r"\s+", " ", t).strip()


def _cleaned_synonyms():
    return {
        metric: sorted({clean_label(s) for s in spec["synonyms"]})
        for metric, spec in METRICS.items()
    }


CLEANED_SYNONYMS = _cleaned_synonyms()


def match_label(label):
    """Return (metric, score) or (None, 0)."""
    t = clean_label(label)
    if len(t) < 3:
        return None, 0
    best, best_score, best_syn = None, 0, ""
    for metric, syns in CLEANED_SYNONYMS.items():
        for syn in syns:
            s = fuzz.token_sort_ratio(t, syn)
            if s > best_score:
                best, best_score, best_syn = metric, s, syn
    if best_score >= MATCH_THRESHOLD:
        # an "Other …" line is a residual, not the total it resembles:
        # "Other current liabilities" must not satisfy current_liabilities
        if t.startswith("other ") and "other" not in best_syn:
            return None, 0
        # directional qualifiers are identity, not noise: "net profit AFTER
        # minority interest" scores 92 against the "... BEFORE minority
        # interest" synonym — a one-token swap the threshold can't separate
        lt, st = set(t.split()), set(best_syn.split())
        if (
            ({"before", "after"} & lt)
            and ({"before", "after"} & st)
            and ("before" in lt) != ("before" in st)
        ):
            return None, 0
        return best, best_score
    return None, 0


def _label_matches(label):
    """All (metric, score) readings of a label. An appositive label —
    "Profit for the year, representing total comprehensive income for the
    year" (Newgen) — names the SAME number twice; both parts are matched
    so one line can satisfy two metrics."""
    t = clean_label(label)
    if " representing " in t:
        parts = [match_label(p) for p in t.split(" representing ")]
        matches = [(m, s) for m, s in parts if m]
        if matches:
            return matches
    m, s = match_label(label)
    return [(m, s)] if m else []


def normalize(raw_items, allowed_metrics=None):
    """RawItems -> best Extraction per canonical metric."""
    by_metric = {}
    for item in raw_items:
        for metric, score in _label_matches(item.label):
            if allowed_metrics is not None and metric not in allowed_metrics:
                continue
            # duplicate labels under both BS sections ("- Trade
            # receivables"): a current-side metric must not take the
            # non-current section's row. A wrong veto degrades to MISSING,
            # never to a wrong value.
            want_side = METRICS[metric].get("side")
            if want_side and getattr(item, "side", None) not in (None, want_side):
                continue
            toks = list(item.values)
            note_stripped = False
            if toks and _looks_like_note_ref(toks[0], toks[1:], item.label):
                toks = toks[1:]
                note_stripped = True
            # second reference column: a Page cross-ref (Note + Page columns,
            # BHEL). Only after a note ref was stripped — the double-column
            # signature — so single-reference statements are untouched.
            if note_stripped and len(toks) >= 2 and _looks_like_page_ref(toks[0], toks[1]):
                toks = toks[1:]
            values = [parse_number(t) for t in toks]
            values = [v for v in values if v is not None]
            if not values:
                continue
            ext = Extraction(
                metric=metric,
                value=values[0],
                raw_label=item.label,
                page=item.page,
                source=item.source,
                score=score,
                extra_values=values[1:],
            )
            prev = by_metric.get(metric)
            if prev is None or score > prev.score:
                by_metric[metric] = ext
    return by_metric


# =============================================================================

# VALIDATOR (from finagent/validator.py)

# =============================================================================

class Status(str, Enum):
    VERIFIED = "VERIFIED"
    PROBABLE = "PROBABLE"
    FLAGGED = "FLAGGED"
    MISSING = "MISSING"
    DERIVED = "DERIVED"  # filled in by the deriver stage, never by extraction


# (name, lhs metrics, rhs metrics, optional rhs metrics)
# meaning sum(lhs) == sum(rhs); optional members join only when extracted
IDENTITY_CHECKS = [
    ("balance_sheet_identity", ["total_assets"], ["total_liabilities", "total_equity"], []),
    ("equity_liabilities_total", ["total_assets"], ["total_equity_and_liabilities"], []),
    ("assets_composition", ["total_assets"], ["current_assets", "non_current_assets"], []),
    (
        "liabilities_composition",
        ["total_liabilities"],
        ["current_liabilities", "non_current_liabilities"],
        [],
    ),
    ("total_income_buildup", ["total_income"], ["revenue", "other_income"], []),
    ("net_profit_buildup", ["profit_before_tax"], ["net_profit", "tax_expense"], []),
    (
        "cash_flow_total",
        ["net_change_in_cash"],
        ["cash_from_operations", "cash_from_investing", "cash_from_financing"],
        ["fx_effect_on_cash"],
    ),
]

# charges presented as a positive figure in Indian reports but as a negative
# line in EU/IFRS layouts ("Income taxes  − 2,785"); identities must hold
# under either convention
SIGN_AMBIGUOUS = {"tax_expense"}

# the same number must appear in two statements: (name, metric_a, metric_b)
CROSS_STATEMENT_CHECKS = [
    ("closing_cash_ties_to_balance_sheet", "cash_and_equivalents", "closing_cash"),
]


def within_tolerance(a, b, rel=0.005, abs_floor=2.0):
    """Rounding-aware equality: reports round to the printed unit, so a sum
    of rounded line items can legitimately differ by a few units."""
    return abs(a - b) <= max(abs_floor, rel * max(abs(a), abs(b)))


@dataclass
class MetricVerdict:
    metric: str
    status: Status
    value: float = None
    sources: list = field(default_factory=list)
    page: int = None
    checks_passed: list = field(default_factory=list)
    checks_failed: list = field(default_factory=list)


class Validator:
    def __init__(self, expected_metrics=None):
        # metric -> list of (value, source, page, label_text)
        self.values = defaultdict(list)
        self.expected = expected_metrics or []

    def add(self, metric, value, source, page=None, label_text=None):
        self.values[metric].append((float(value), source, page, label_text))

    def _consensus(self, metric):
        """Median of proposals — one wild value can't drag it."""
        vals = self.values.get(metric)
        return median(v for v, *_ in vals) if vals else None

    def validate(self):
        consensus = {m: self._consensus(m) for m in self.values}
        passed, failed = defaultdict(list), defaultdict(list)

        def totals(metrics):
            """All defensible sums: a negative sign-ambiguous charge also
            contributes its flipped value as an alternative."""
            if any(consensus.get(m) is None for m in metrics):
                return None
            sums = [0.0]
            for m in metrics:
                v = consensus[m]
                if m in SIGN_AMBIGUOUS and v < 0:
                    sums = [s + v for s in sums] + [s - v for s in sums]
                else:
                    sums = [s + v for s in sums]
            return sums

        # extractors that disagree flag their metric
        for m, obs in self.values.items():
            if len(obs) > 1:
                agree = all(within_tolerance(v, consensus[m]) for v, *_ in obs)
                (passed if agree else failed)[m].append("cross_extractor_voting")

        # accounting identities. Optional members (fx effect) sit INSIDE the
        # sum in some layouts (BMW) but OUTSIDE it, in the opening->closing
        # reconciliation, in others (Airtel "(a+b+c)") — the identity passes
        # if either reading ties. The with-fx reading is tried first so fx
        # keeps its VERIFIED credit where it genuinely belongs to the sum.
        for name, lhs, rhs, opt in IDENTITY_CHECKS:
            present = [m for m in opt if consensus.get(m) is not None]
            ls = totals(lhs)
            if ls is None or totals(rhs) is None:
                continue
            variants = ([rhs + present] if present else []) + [rhs]
            ok_members = next(
                (
                    members
                    for members in variants
                    if any(within_tolerance(l, r) for l in ls for r in totals(members))
                ),
                None,
            )
            detail = f"{name} ({ls[0]:,.1f} vs {totals(rhs)[0]:,.1f})"
            if ok_members is not None:
                for m in lhs + ok_members:
                    passed[m].append(detail)
            else:
                for m in lhs + rhs + present:
                    failed[m].append(detail)

        # cross-statement ties
        for name, a, b in CROSS_STATEMENT_CHECKS:
            va, vb = consensus.get(a), consensus.get(b)
            if va is None or vb is None:
                continue
            detail = f"{name} ({va:,.1f} vs {vb:,.1f})"
            for m in (a, b):
                (passed if within_tolerance(va, vb) else failed)[m].append(detail)

        verdicts = {}
        for m, obs in self.values.items():
            value = consensus[m]
            best = min(obs, key=lambda o: abs(o[0] - value))
            status = (
                Status.FLAGGED if failed[m] else Status.VERIFIED if passed[m] else Status.PROBABLE
            )
            verdicts[m] = MetricVerdict(
                metric=m,
                status=status,
                value=value,
                sources=sorted({s for _, s, _, _ in obs}),
                page=best[2],
                checks_passed=passed[m],
                checks_failed=failed[m],
            )
        for m in self.expected:
            if m not in verdicts:
                verdicts[m] = MetricVerdict(metric=m, status=Status.MISSING)
        return ValidationReport(verdicts)


class ValidationReport:
    def __init__(self, verdicts):
        self.verdicts = verdicts

    def by_status(self, status):
        return [v for v in self.verdicts.values() if v.status == status]

    def print_summary(self):
        icon = {
            Status.VERIFIED: "[OK]",
            Status.PROBABLE: "[?] ",
            Status.FLAGGED: "[!!]",
            Status.MISSING: "[--]",
            Status.DERIVED: "[=>]",
        }
        print("=" * 72)
        for status in (
            Status.FLAGGED,
            Status.MISSING,
            Status.DERIVED,
            Status.VERIFIED,
            Status.PROBABLE,
        ):
            for v in sorted(self.by_status(status), key=lambda x: x.metric):
                val = f"{v.value:,.1f}" if v.value is not None else "-"
                src = ",".join(v.sources) if v.sources else "-"
                print(f"{icon[v.status]} {v.metric:<32} {val:>15}  [{src}]")
                for c in v.checks_failed:
                    print(f"      failed: {c}")
        print("-" * 72)
        c = {s: len(self.by_status(s)) for s in Status}
        print(
            f"VERIFIED: {c[Status.VERIFIED]}   PROBABLE: {c[Status.PROBABLE]}   "
            f"FLAGGED: {c[Status.FLAGGED]}   DERIVED: {c[Status.DERIVED]}   "
            f"MISSING: {c[Status.MISSING]}"
        )
        print("=" * 72)

    def to_dict(self):
        return {
            m: {
                "status": v.status.value,
                "value": v.value,
                "sources": v.sources,
                "page": v.page,
                "checks_passed": v.checks_passed,
                "checks_failed": v.checks_failed,
            }
            for m, v in self.verdicts.items()
        }


# =============================================================================

# DERIVER (from finagent/deriver.py)

# =============================================================================

# target <- [(input_metric, sign), ...]
DERIVATIONS = [
    ("total_liabilities", [("total_equity_and_liabilities", 1), ("total_equity", -1)]),
    ("total_liabilities", [("total_assets", 1), ("total_equity", -1)]),
    ("total_equity", [("total_equity_and_liabilities", 1), ("total_liabilities", -1)]),
    ("total_equity", [("total_assets", 1), ("total_liabilities", -1)]),
    ("current_assets", [("total_assets", 1), ("non_current_assets", -1)]),
    ("non_current_assets", [("total_assets", 1), ("current_assets", -1)]),
    ("current_liabilities", [("total_liabilities", 1), ("non_current_liabilities", -1)]),
    ("non_current_liabilities", [("total_liabilities", 1), ("current_liabilities", -1)]),
    # the balance-sheet identity itself: both sides are the same number
    ("total_equity_and_liabilities", [("total_assets", 1)]),
    ("total_assets", [("total_equity_and_liabilities", 1)]),
    ("total_income", [("revenue", 1), ("other_income", 1)]),
    # P&L identity: total income - total expenses = profit before tax. Function-
    # of-expense statements (Newgen) and some IFRS layouts never print a "Total
    # expenses" line; the identity pins it exactly from income and PBT.
    ("total_expenses", [("total_income", 1), ("profit_before_tax", -1)]),
    ("total_expenses", [("revenue", 1), ("other_income", 1), ("profit_before_tax", -1)]),
    # some P&Ls print tax only as Current/Deferred sub-lines with no total
    # (Airtel). Sign risk, documented: an EU-convention file (tax printed
    # negative) with tax MISSING would derive the positive magnitude — no
    # such case in corpus, and the positive-only guard logs any oddity.
    ("tax_expense", [("profit_before_tax", 1), ("net_profit", -1)]),
]

_TRUSTED = {Status.VERIFIED, Status.PROBABLE}


def _expr(formula, inputs):
    parts = [
        f"{'-' if s < 0 else '+'} {m}[{i.status.value[0]}]" for (m, s), i in zip(formula, inputs)
    ]
    return " ".join(parts).lstrip("+ ")


def derive(report):
    """Fill MISSING verdicts in a ValidationReport with DERIVED ones."""
    verdicts = report.verdicts
    by_target = {}
    for target, formula in DERIVATIONS:
        by_target.setdefault(target, []).append(formula)
    for target, formulas in by_target.items():
        v = verdicts.get(target)
        if v is None or v.status != Status.MISSING:
            continue
        # all computable formulas compete; the one resting on the fewest
        # unproven (non-VERIFIED) inputs wins — code order only breaks ties
        candidates = []
        for order, formula in enumerate(formulas):
            inputs = [verdicts.get(m) for m, _ in formula]
            if any(i is None or i.status not in _TRUSTED for i in inputs):
                continue
            unproven = sum(1 for i in inputs if i.status != Status.VERIFIED)
            candidates.append((unproven, order, formula, inputs))
        if not candidates:
            continue
        _, _, formula, inputs = min(candidates, key=lambda c: c[:2])
        value = sum(sign * i.value for (_, sign), i in zip(formula, inputs))
        if value <= 0:
            # a non-positive result means an upstream extraction is corrupt —
            # keep MISSING but leave the diagnostic, don't discard silently
            v.checks_failed.append(f"derivation discarded (non-positive): {_expr(formula, inputs)}")
            continue
        verdicts[target] = MetricVerdict(
            metric=target,
            status=Status.DERIVED,
            value=value,
            sources=["derived"],
            page=inputs[0].page,
            checks_passed=[f"derived: {_expr(formula, inputs)}"],
        )
    return report


# =============================================================================

# WRITER (from finagent/writer.py)

# =============================================================================

def _clean(value):
    """Strip control/illegal chars openpyxl refuses to write to a cell.
    PDF text layers occasionally carry stray control bytes inside labels
    (Bajaj's "Other operating income" arrived as a control-joined token);
    those would crash the whole write. Non-strings pass through untouched."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


STATUS_FILL = {
    "VERIFIED": PatternFill("solid", fgColor="C6EFCE"),  # green
    "PROBABLE": PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "FLAGGED": PatternFill("solid", fgColor="FFC7CE"),  # red
    "MISSING": PatternFill("solid", fgColor="D9D9D9"),  # grey
    "DERIVED": PatternFill("solid", fgColor="BDD7EE"),  # blue: computed,
}  # not extracted
STATEMENT_NAMES = {"PL": "Profit & Loss", "BS": "Balance Sheet", "CF": "Cash Flow"}
HEADERS = [
    "Statement",
    "Metric",
    "Value",
    "Status",
    "Page",
    "Matched label",
    "Sources",
    "Checks passed",
    "Checks failed",
]


def _fill_sheet(ws, report_dict, extractions, meta):
    """Render one basis (report_dict + extractions) onto a worksheet."""
    if meta:
        ws.append([meta])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])

    ws.append(HEADERS)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = Font(bold=True)

    extractions = extractions or {}
    for metric, spec in METRICS.items():
        v = report_dict.get(
            metric,
            {
                "status": "MISSING",
                "value": None,
                "sources": [],
                "page": None,
                "checks_passed": [],
                "checks_failed": [],
            },
        )
        ext = extractions.get(metric)
        page = v.get("page")
        ws.append(
            [
                _clean(x)
                for x in (
                    STATEMENT_NAMES[spec["statement"]],
                    metric,
                    v["value"],
                    v["status"],
                    (page + 1) if page is not None else None,  # 1-based for humans
                    ext.raw_label if ext else "",
                    ", ".join(v.get("sources", [])),
                    "; ".join(v.get("checks_passed", [])),
                    "; ".join(v.get("checks_failed", [])),
                )
            ]
        )
        ws.cell(row=ws.max_row, column=4).fill = STATUS_FILL[v["status"]]

    widths = [14, 28, 16, 11, 6, 45, 18, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


SIDE_BY_SIDE_HEADERS = [
    "Statement",
    "Metric",
    "Consolidated Value",
    "Standalone Value",
    "Difference (Consol - Standalone)",
    "Consolidated Status",
    "Standalone Status",
    "Consolidated Page",
    "Standalone Page",
]


def _fill_side_by_side_sheet(ws, comp_dict, meta):
    """Render side-by-side comparison of Consolidated vs Standalone."""
    if meta:
        ws.append([meta])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])

    ws.append(SIDE_BY_SIDE_HEADERS)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = Font(bold=True)

    for metric, spec in METRICS.items():
        v = comp_dict.get(metric, {})
        c_val = v.get("consolidated_value")
        s_val = v.get("standalone_value")
        diff = v.get("difference")
        c_status = v.get("consolidated_status", "MISSING")
        s_status = v.get("standalone_status", "MISSING")
        c_page = v.get("consolidated_page")
        s_page = v.get("standalone_page")

        ws.append(
            [
                _clean(x)
                for x in (
                    STATEMENT_NAMES[spec["statement"]],
                    metric,
                    c_val,
                    s_val,
                    diff,
                    c_status,
                    s_status,
                    (c_page + 1) if c_page is not None else None,
                    (s_page + 1) if s_page is not None else None,
                )
            ]
        )
        ws.cell(row=ws.max_row, column=6).fill = STATUS_FILL.get(c_status, STATUS_FILL["MISSING"])
        ws.cell(row=ws.max_row, column=7).fill = STATUS_FILL.get(s_status, STATUS_FILL["MISSING"])

    widths = [14, 28, 20, 18, 30, 20, 18, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_excel(sheets, out_path, extractions=None, meta=None):
    """Write one worksheet per basis.

    sheets: list of (sheet_name, report_dict, extractions). For backward
    compatibility a single (report_dict, extractions=...) call is still
    accepted — it becomes a one-sheet "Metrics" workbook.
    """
    # back-compat: write_excel(report_dict, path, extractions=..., meta=...)
    if isinstance(sheets, dict):
        sheets = [("Metrics", sheets, extractions)]

    wb = Workbook()
    wb.remove(wb.active)  # we add named sheets explicitly
    for name, report_dict, ext in sheets:
        ws = wb.create_sheet(title=name[:31])  # Excel caps sheet names at 31
        if name in ("Side-by-Side", "Comparison"):
            _fill_side_by_side_sheet(ws, report_dict, meta)
        else:
            _fill_sheet(ws, report_dict, ext, meta)

    wb.save(out_path)
    return out_path


# =============================================================================

# TRACING (from finagent/tracing.py)

# =============================================================================

DEFAULT_SITE_ID = "e3751324-7684-484d-8dd5-b7266df80bcb"
DEFAULT_AGENT_NAME = "financial-pdf-extraction-agent"

_INITIALIZED = False
_TRODO_AVAILABLE = False

try:

    _TRODO_AVAILABLE = True
except ImportError:
    trodo = None


class DummyRun:
    """Fallback run object if Trodo is not installed."""

    def set_input(self, data: Any) -> None:
        pass

    def set_output(self, data: Any) -> None:
        pass

    def set_metadata(self, data: dict[str, Any]) -> None:
        pass

    def set_error_summary(self, summary: str) -> None:
        pass


class DummySpan:
    """Fallback span object if Trodo is not installed."""

    def set_input(self, data: Any) -> None:
        pass

    def set_output(self, data: Any) -> None:
        pass

    def set_metadata(self, data: dict[str, Any]) -> None:
        pass


def init_tracing(site_id: str | None = None) -> bool:
    """Initialize Trodo tracing once per runtime process."""
    global _INITIALIZED
    if not _TRODO_AVAILABLE or trodo is None:
        return False
    if _INITIALIZED:
        return True

    effective_site_id = site_id or os.getenv("TRODO_SITE_ID", DEFAULT_SITE_ID)
    try:
        trodo.init(site_id=effective_site_id)
        _INITIALIZED = True
        return True
    except Exception as err:  # noqa: BLE001
        print(f"[finagent.tracing] Failed to initialize Trodo tracing: {err}")
        return False


@contextmanager
def wrap_agent(
    name: str = DEFAULT_AGENT_NAME,
    distinct_id: str | None = None,
    metadata: dict[str, Any] | None = None,
):
    """Context manager wrapping an entire agent execution run in Trodo."""
    init_tracing()

    effective_distinct_id = distinct_id or os.getenv("TRODO_DISTINCT_ID", "cli-user")
    if _INITIALIZED and trodo is not None:
        kwargs = {"distinct_id": effective_distinct_id}
        if metadata:
            kwargs["metadata"] = metadata
        with trodo.wrap_agent(name, **kwargs) as run:
            yield run
    else:
        yield DummyRun()


@contextmanager
def trace_span(name: str, kind: str = "trace"):
    """Context manager for child spans inside an agent run."""
    if _INITIALIZED and trodo is not None:
        with trodo.span(name, kind=kind) as s:
            yield s
    else:
        yield DummySpan()


def flush_tracing() -> None:
    """Flush pending telemetry batches to Trodo."""
    if _INITIALIZED and trodo is not None:
        with contextlib.suppress(Exception):
            trodo.flush()


# =============================================================================

# PIPELINE (from finagent/pipeline.py)

# =============================================================================

def _extract_basis(pdf_path, locations, log, label):
    """Run extract -> normalize -> validate -> derive for one set of statement
    locations (one basis). Returns (report, {metric: Extraction})."""

    v = Validator(expected_metrics=EXPECTED_METRICS)
    all_extractions = {}
    for code, loc in locations.items():
        if not loc.page_indices:
            continue
        raw = geometric.extract(
            pdf_path,
            loc.page_indices,
            cue_pats=locator.STATEMENT_SIGNATURES[code][1],
            want_basis=label,
        )
        extractions = normalizer.normalize(raw, allowed_metrics=set(metrics_for_statement(code)))
        log(f"[extract:{label}] {code}: {len(raw)} lines -> {len(extractions)} metrics matched")
        for metric, ext in extractions.items():
            if METRICS[metric]["statement"] != code and metric in all_extractions:
                continue
            all_extractions[metric] = ext
            v.add(metric, ext.value, source=ext.source, page=ext.page, label_text=ext.raw_label)

    report = derive(v.validate())
    return report, all_extractions


def _sheet_name(basis):
    return {"consolidated": "Consolidated", "standalone": "Standalone"}.get(basis, "Financials")


def run(pdf_path, out_path=None, verbose=True):
    pdf_path = Path(pdf_path)
    t0 = time.time()

    def log(msg):
        if verbose:
            print(msg)

    with wrap_agent("financial-pdf-extraction-agent") as agent_run:
        agent_run.set_input(
            {
                "pdf_path": str(pdf_path),
                "out_path": str(out_path) if out_path is not None else None,
            }
        )

        # 1. profile
        with trace_span("pdf_profile", kind="trace"):
            doc = profiler.profile(pdf_path)
            log(f"[profile] {doc.summary()}")

        # 1b. unit & period anchor engine
        with trace_span("unit_period_detection", kind="trace"):
            sample_text = " ".join(p.text for p in doc.pages[:15])
            unit_info = unit_detector.detect_unit(sample_text)
            period_info = unit_detector.detect_periods(sample_text)
            log(
                f"[unit_anchor] Scale: {unit_info.unit_name} ({unit_info.currency}) | mult={unit_info.multiplier}"
            )
            log(
                f"[period_anchor] Periods: current='{period_info.current_period}', prior='{period_info.prior_period}'"
            )

        # 3. locate statement pages — PRIMARY (prefers consolidated) plus the
        # standalone/consolidated counterpart, so both are extracted separately.
        with trace_span("locate_pages", kind="trace"):
            primary = locator.locate(doc)
            alternate = locator.locate_alternate(doc, primary)
            for code, loc in primary.items():
                pages_1based = [i + 1 for i in loc.page_indices]
                log(
                    f"[locate] {code}: pages {pages_1based} basis={loc.basis} score={loc.score:.1f}"
                )
                alt = alternate.get(code)
                if alt and alt.page_indices:
                    log(
                        f"[locate]   alt {code}: pages {[i + 1 for i in alt.page_indices]} "
                        f"basis={alt.basis} score={alt.score:.1f}"
                    )

        # 4-6b. extract + validate + derive, once per basis. The primary report is
        # the one returned (backward-compatible with the golden/benchmark harness);
        # the alternate is the standalone counterpart, shown on its own sheet.
        primary_basis = primary["BS"].basis if primary.get("BS") else "unknown"
        primary_label = _sheet_name(primary_basis).lower()
        with trace_span(f"extract_basis_{primary_label}", kind="tool"):
            report, primary_ext = _extract_basis(pdf_path, primary, log, primary_label)
        if verbose:
            report.print_summary()

        sheets = [(_sheet_name(primary_basis), report.to_dict(), primary_ext)]
        if locator.has_pages(alternate):
            alt_basis = (
                alternate["BS"].basis
                if alternate.get("BS") and alternate["BS"].page_indices
                else "standalone"
            )
            with trace_span(f"extract_basis_{alt_basis}", kind="tool"):
                alt_report, alt_ext = _extract_basis(pdf_path, alternate, log, alt_basis)
            sheets.append((_sheet_name(alt_basis), alt_report.to_dict(), alt_ext))

            # Build Side-by-Side Comparison dict
            comp_dict = {}
            primary_dict = report.to_dict()
            alt_dict = alt_report.to_dict()

            consol_d = primary_dict if primary_label == "consolidated" else alt_dict
            stand_d = alt_dict if primary_label == "consolidated" else primary_dict

            for metric in METRICS:
                c_v = consol_d.get(metric, {})
                s_v = stand_d.get(metric, {})
                c_val = c_v.get("value")
                s_val = s_v.get("value")
                diff = (c_val - s_val) if (c_val is not None and s_val is not None) else None

                comp_dict[metric] = {
                    "consolidated_value": c_val,
                    "standalone_value": s_val,
                    "difference": diff,
                    "consolidated_status": c_v.get("status", "MISSING"),
                    "standalone_status": s_v.get("status", "MISSING"),
                    "consolidated_page": c_v.get("page"),
                    "standalone_page": s_v.get("page"),
                }

            sheets.insert(0, ("Side-by-Side", comp_dict, None))

        # 7. write — one sheet per basis
        if out_path is None:
            out_path = pdf_path.with_suffix("").name + "_metrics.xlsx"
            out_path = Path("output") / out_path
            out_path.parent.mkdir(exist_ok=True)

        with trace_span("write_excel", kind="tool"):
            write_excel(
                sheets,
                out_path,
                meta=f"{pdf_path.name} - extracted {time.strftime('%Y-%m-%d %H:%M')}",
            )
        log(f"[write] {out_path}  ({time.time() - t0:.1f}s)")

        agent_run.set_output(
            {
                "status": "success",
                "pdf_name": pdf_path.name,
                "out_path": str(out_path),
                "primary_basis": primary_basis,
                "sheets": [s[0] for s in sheets],
                "metrics_extracted": len(primary_ext),
                "execution_time_seconds": round(time.time() - t0, 2),
            }
        )
        flush_tracing()
        return report



if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python finagent_single.py <pdf_path> [out_excel_path]")
        sys.exit(1)
    run(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
