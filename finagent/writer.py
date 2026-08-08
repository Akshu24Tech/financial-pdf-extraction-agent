"""Stage 7: Excel output with receipts.

One row per metric: value, validation status, page citation, the exact
label text it was matched from, sources, and which checks passed/failed.
"""

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill

from .schema import METRICS


def _clean(value):
    """Strip control/illegal chars openpyxl refuses to write to a cell.
    PDF text layers occasionally carry stray control bytes inside labels
    (Bajaj's "Other operating income" arrived as a control-joined token);
    those would crash the whole write. Non-strings pass through untouched."""
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub("", value)
    return value


STATUS_FILL = {
    "VERIFIED": PatternFill("solid", fgColor="C6EFCE"),  # green
    "PROBABLE": PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "FLAGGED": PatternFill("solid", fgColor="FFC7CE"),  # red
    "MISSING": PatternFill("solid", fgColor="D9D9D9"),  # grey
    "DERIVED": PatternFill("solid", fgColor="BDD7EE"),  # blue: computed,
}  # not extracted
STATEMENT_NAMES = {"PL": "Profit & Loss", "BS": "Balance Sheet", "CF": "Cash Flow"}
HEADERS = [
    "Statement",
    "Metric",
    "Value",
    "Status",
    "Page",
    "Matched label",
    "Sources",
    "Checks passed",
    "Checks failed",
]


def _fill_sheet(ws, report_dict, extractions, meta):
    """Render one basis (report_dict + extractions) onto a worksheet."""
    if meta:
        ws.append([meta])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])

    ws.append(HEADERS)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = Font(bold=True)

    extractions = extractions or {}
    for metric, spec in METRICS.items():
        v = report_dict.get(
            metric,
            {
                "status": "MISSING",
                "value": None,
                "sources": [],
                "page": None,
                "checks_passed": [],
                "checks_failed": [],
            },
        )
        ext = extractions.get(metric)
        page = v.get("page")
        ws.append(
            [
                _clean(x)
                for x in (
                    STATEMENT_NAMES[spec["statement"]],
                    metric,
                    v["value"],
                    v["status"],
                    (page + 1) if page is not None else None,  # 1-based for humans
                    ext.raw_label if ext else "",
                    ", ".join(v.get("sources", [])),
                    "; ".join(v.get("checks_passed", [])),
                    "; ".join(v.get("checks_failed", [])),
                )
            ]
        )
        ws.cell(row=ws.max_row, column=4).fill = STATUS_FILL[v["status"]]

    widths = [14, 28, 16, 11, 6, 45, 18, 40, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


SIDE_BY_SIDE_HEADERS = [
    "Statement",
    "Metric",
    "Consolidated Value",
    "Standalone Value",
    "Difference (Consol - Standalone)",
    "Consolidated Status",
    "Standalone Status",
    "Consolidated Page",
    "Standalone Page",
]


def _fill_side_by_side_sheet(ws, comp_dict, meta):
    """Render side-by-side comparison of Consolidated vs Standalone."""
    if meta:
        ws.append([meta])
        ws["A1"].font = Font(bold=True, size=12)
        ws.append([])

    ws.append(SIDE_BY_SIDE_HEADERS)
    header_row = ws.max_row
    for c in ws[header_row]:
        c.font = Font(bold=True)

    for metric, spec in METRICS.items():
        v = comp_dict.get(metric, {})
        c_val = v.get("consolidated_value")
        s_val = v.get("standalone_value")
        diff = v.get("difference")
        c_status = v.get("consolidated_status", "MISSING")
        s_status = v.get("standalone_status", "MISSING")
        c_page = v.get("consolidated_page")
        s_page = v.get("standalone_page")

        ws.append(
            [
                _clean(x)
                for x in (
                    STATEMENT_NAMES[spec["statement"]],
                    metric,
                    c_val,
                    s_val,
                    diff,
                    c_status,
                    s_status,
                    (c_page + 1) if c_page is not None else None,
                    (s_page + 1) if s_page is not None else None,
                )
            ]
        )
        ws.cell(row=ws.max_row, column=6).fill = STATUS_FILL.get(c_status, STATUS_FILL["MISSING"])
        ws.cell(row=ws.max_row, column=7).fill = STATUS_FILL.get(s_status, STATUS_FILL["MISSING"])

    widths = [14, 28, 20, 18, 30, 20, 18, 18, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=header_row, column=i).column_letter].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


def write_excel(sheets, out_path, extractions=None, meta=None):
    """Write one worksheet per basis.

    sheets: list of (sheet_name, report_dict, extractions). For backward
    compatibility a single (report_dict, extractions=...) call is still
    accepted — it becomes a one-sheet "Metrics" workbook.
    """
    # back-compat: write_excel(report_dict, path, extractions=..., meta=...)
    if isinstance(sheets, dict):
        sheets = [("Metrics", sheets, extractions)]

    wb = Workbook()
    wb.remove(wb.active)  # we add named sheets explicitly
    for name, report_dict, ext in sheets:
        ws = wb.create_sheet(title=name[:31])  # Excel caps sheet names at 31
        if name in ("Side-by-Side", "Comparison"):
            _fill_side_by_side_sheet(ws, report_dict, meta)
        else:
            _fill_sheet(ws, report_dict, ext, meta)

    wb.save(out_path)
    return out_path
